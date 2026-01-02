from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, session, send_file
import os
from datetime import datetime
import pandas as pd
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import calendar
from datetime import timedelta
from collections import Counter

# Importar infraestructura
from src.infrastructure.mysql_connection import MySQLConnection
from src.infrastructure.repositories_mysql import (
    EmpresaRepositoryMySQL,
    EmpleadoRepositoryMySQL,
    AsistenciaRepositoryMySQL,
    HorarioEstandarRepositoryMySQL,
    EscaneoTrackingRepositoryMySQL,
    AdministradorRepository
)

# Importar use cases
from src.use_cases.register_employee import RegisterEmployeeUseCase
from src.use_cases.mark_attendance import MarkAttendanceUseCase
from src.use_cases.list_companies import ListCompaniesUseCase
from src.use_cases.get_report import GetReportUseCase, minutos_a_hhmm

# Importar QR generator
from src.infrastructure.qr_generator import QRGenerator

app = Flask(__name__)
app.secret_key = os.getenv('SECRET_KEY') or 'clave-secreta-temporal-desarrollo-cambiar-en-produccion'

# Configuración de base de datos
db_connection = MySQLConnection()
# Inicializar repositorios
empresa_repo = EmpresaRepositoryMySQL(db_connection)
empleado_repo = EmpleadoRepositoryMySQL(db_connection)
asistencia_repo = AsistenciaRepositoryMySQL(db_connection)
horario_repo = HorarioEstandarRepositoryMySQL(db_connection)
escaneo_repo = EscaneoTrackingRepositoryMySQL(db_connection)

# Inicializar use cases
register_employee_use_case = RegisterEmployeeUseCase(empleado_repo)
mark_attendance_use_case = MarkAttendanceUseCase(empleado_repo, asistencia_repo, horario_repo, escaneo_repo)
list_companies_use_case = ListCompaniesUseCase(empresa_repo,)
get_report_use_case = GetReportUseCase(empleado_repo, asistencia_repo, empresa_repo)

# Inicializar QR generator
qr_generator = QRGenerator()

def obtener_nombre_mes(numero_mes):
    """Obtiene el nombre del mes por su número"""
    meses = [
        '', 'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
        'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre'
    ]
    return meses[numero_mes] if 1 <= numero_mes <= 12 else ''

# Rutas principales
@app.route('/')
def index():
    return render_template('scan.html')

# Rutas de administración
@app.route('/admin')
def admin_dashboard():
    if not session.get('admin_logged_in'):
        return redirect(url_for('admin_login'))
    
    empresas = list_companies_use_case.execute()
    return render_template('admin_dashboard.html', empresas=empresas)

@app.route('/admin/login', methods=['GET', 'POST'])
def admin_login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        
        admin_repo = AdministradorRepository(db_connection)
        administrador = admin_repo.get_by_username(username)
        
        if administrador and admin_repo.verify_password(administrador['password_hash'], password):
            session['admin_logged_in'] = True
            session['admin_id'] = administrador['id']
            session['admin_nombre'] = administrador['nombre']
            session['admin_rol'] = administrador['rol']
            flash('Inicio de sesión exitoso', 'success')
            return redirect(url_for('admin_dashboard'))
        else:
            flash('Credenciales inválidas', 'error')
    
    return render_template('login.html')

@app.route('/admin/logout')
def admin_logout():
    session.clear()
    return redirect(url_for('index'))

@app.route('/admin/add_employee', methods=['GET', 'POST'])
def admin_add_employee():
    if not session.get('admin_logged_in'):
        return redirect(url_for('admin_login'))
    
    if request.method == 'POST':
        try:
            nombre = request.form['nombre']
            empresa_id = int(request.form['empresa_id'])
            dni = request.form['dni']
            telefono = request.form.get('telefono', '')
            correo = request.form.get('correo', '')
            
            empleado = register_employee_use_case.execute(
                nombre=nombre,
                empresa_id=empresa_id,
                dni=dni,
                telefono=telefono,
                correo=correo
            )
            
            empresa = empresa_repo.get_by_id(empresa_id)
            if empresa:
                qr_path = qr_generator.generate_employee_qr(empleado.id, empresa.codigo_empresa)
                if qr_path:
                    flash('Empleado registrado con éxito. Código QR generado.', 'success')
                else:
                    flash('Empleado registrado, pero hubo un error generando el QR.', 'warning')
            else:
                flash('Empleado registrado con éxito.', 'success')
            
            return redirect(url_for('admin_add_employee'))
            
        except Exception as e:
            flash(f'Error registrando empleado: {str(e)}', 'error')
    
    empresas = list_companies_use_case.execute()
    return render_template('admin_add_employee.html', empresas=empresas)

@app.route('/admin/employees')
def admin_list_employees():
    if not session.get('admin_logged_in'):
        return redirect(url_for('admin_login'))
    
    empresa_id = request.args.get('empresa_id', type=int)
    
    if empresa_id:
        empleados = empleado_repo.get_by_empresa_id(empresa_id)
        empresa = empresa_repo.get_by_id(empresa_id)
    else:
        empleados = empleado_repo.get_all()
        empresa = None
    
    empresas = list_companies_use_case.execute()
    return render_template('admin_list_employees.html', 
                         empleados=empleados, 
                         empresas=empresas, 
                         empresa_seleccionada=empresa)

@app.route('/admin/edit_employee/<int:empleado_id>', methods=['GET', 'POST'])
def edit_employee(empleado_id):
    if not session.get('admin_logged_in'):
        return redirect(url_for('admin_login'))
    
    empleado = empleado_repo.get_by_id(empleado_id)
    if not empleado:
        flash('Empleado no encontrado', 'error')
        return redirect(url_for('admin_list_employees'))
    
    if request.method == 'POST':
        try:
            empleado.nombre = request.form['nombre']
            empleado.empresa_id = int(request.form['empresa_id'])
            empleado.dni = request.form['dni']
            empleado.telefono = request.form.get('telefono', '')
            empleado.correo = request.form.get('correo', '')
            
            empleado_repo.update(empleado)
            flash('Empleado actualizado con éxito', 'success')
            return redirect(url_for('admin_list_employees'))
            
        except Exception as e:
            flash(f'Error actualizando empleado: {str(e)}', 'error')
    
    empresas = list_companies_use_case.execute()
    return render_template('admin_edit_employee.html', empleado=empleado, empresas=empresas)

@app.route('/admin/toggle_employee/<int:empleado_id>', methods=['POST'])
def toggle_employee(empleado_id):
    if not session.get('admin_logged_in'):
        return jsonify({'success': False, 'message': 'No autorizado'}), 401
    
    try:
        empleado = empleado_repo.get_by_id(empleado_id)
        if not empleado:
            return jsonify({'success': False, 'message': 'Empleado no encontrado'}), 404
        
        empleado.activo = not empleado.activo
        empleado_repo.update(empleado)
        
        estado = "activado" if empleado.activo else "desactivado"
        return jsonify({
            'success': True, 
            'message': f'Empleado {estado} con éxito',
            'activo': empleado.activo
        })
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'}), 500

@app.route('/admin/delete_employee/<int:empleado_id>', methods=['POST'])
def delete_employee(empleado_id):
    if not session.get('admin_logged_in'):
        return jsonify({'success': False, 'message': 'No autorizado'}), 401
    
    try:
        empleado = empleado_repo.get_by_id(empleado_id)
        if not empleado:
            return jsonify({'success': False, 'message': 'Empleado no encontrado'}), 404
        
        nombre_empleado = empleado.nombre
        
        empleado_repo.delete(empleado_id)
        
        return jsonify({
            'success': True, 
            'message': f'Empleado {nombre_empleado} eliminado permanentemente de la base de datos'
        })
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'}), 500

@app.route('/admin/download_qr/<int:empleado_id>')
def download_qr(empleado_id):
    try:
        empleado = empleado_repo.get_by_id(empleado_id)
        if not empleado:
            flash('Empleado no encontrado', 'error')
            return redirect(url_for('admin_list_employees'))
        
        empresa = empresa_repo.get_by_id(empleado.empresa_id)
        if not empresa:
            flash('Empresa no encontrada', 'error')
            return redirect(url_for('admin_list_employees'))
        
        qr_path = qr_generator.generate_employee_qr(empleado.id, empresa.codigo_empresa)
        
        if qr_path and os.path.exists(qr_path):
            return send_file(
                qr_path,
                mimetype='image/png',
                as_attachment=True,
                download_name=f'qr_empleado_{empleado_id}_{empresa.codigo_empresa}.png'
            )
        else:
            flash('Error generando código QR para descarga', 'error')
            return redirect(url_for('admin_list_employees'))
            
    except Exception as e:
        flash(f'Error descargando QR: {str(e)}', 'error')
        return redirect(url_for('admin_list_employees'))

@app.route('/scan')
def scan_qr():
    return render_template('scan.html')

@app.route('/api/scan', methods=['POST'])
def api_scan_qr():
    try:
        data = request.get_json()
        codigo_qr = data.get('codigo_qr', '')
        ip_address = request.remote_addr
        
        resultado = mark_attendance_use_case.execute(codigo_qr, ip_address)
        
        return jsonify(resultado)
        
    except Exception as e:
        return jsonify({
            "status": "error",
            "message": f"Error procesando escaneo: {str(e)}",
            "data": None
        })

@app.route('/reports')
def reports():
     if not session.get('admin_logged_in'):
        flash('Debes iniciar sesión para acceder a los reportes', 'error')
        return redirect(url_for('admin_login'))
     
     empresas = list_companies_use_case.execute()
     return render_template('report.html', empresas=empresas)

@app.route('/api/reports/monthly')
def api_monthly_report():
    try:
        empresa_id = request.args.get('empresa_id', type=int)
        mes = request.args.get('mes', type=int, default=datetime.now().month)
        anio = request.args.get('anio', type=int, default=datetime.now().year)
        
        if not empresa_id:
            return jsonify({"error": "Empresa ID requerido"}), 400
        
        reporte = get_report_use_case.execute_monthly_report(empresa_id, mes, anio)
        return jsonify(reporte)
        
    except Exception as e:
        return jsonify({"error": f"Error generando reporte: {str(e)}"}), 500

@app.route('/api/reports/employee/<int:empleado_id>')
def api_employee_report(empleado_id):
    try:    
        mes = request.args.get('mes', type=int, default=datetime.now().month)
        anio = request.args.get('anio', type=int, default=datetime.now().year)
        
        reporte = get_report_use_case.execute_employee_detail_report(empleado_id, mes, anio)
        return jsonify(reporte)
        
    except Exception as e:
        return jsonify({"error": f"Error generando reporte: {str(e)}"}), 500
    
@app.route('/api/empleados')
def api_get_empleados():
    """Obtiene empleados de una empresa"""
    try:
        empresa_id = request.args.get('empresa_id', type=int)
        
        if not empresa_id:
            return jsonify({"error": "empresa_id requerido"}), 400
        
        empleados = empleado_repo.get_by_empresa_id(empresa_id)
        
        # Convertir a diccionario
        empleados_data = []
        for emp in empleados:
            empleados_data.append({
                "id": emp.id,
                "nombre": emp.nombre,
                "dni": emp.dni,
                "empresa_id": emp.empresa_id,
                "activo": emp.activo
            })
        
        return jsonify(empleados_data)
        
    except Exception as e:
        print(f"Error en api_get_empleados: {e}")
        return jsonify({"error": str(e)}), 500
    

@app.route('/api/asistencias/<int:empleado_id>')
def api_get_asistencia_empleado(empleado_id):
    """Obtiene asistencia de un empleado en una fecha específica"""
    try:
        fecha = request.args.get('fecha')
        
        if not fecha:
            return jsonify({"error": "Fecha requerida"}), 400
        
        asistencia = asistencia_repo.get_by_empleado_and_fecha(empleado_id, fecha)
        
        if asistencia:
            return jsonify({
                "fecha": str(asistencia.fecha),
                "entrada_manana_real": str(asistencia.entrada_manana_real) if asistencia.entrada_manana_real else None,
                "salida_manana_real": str(asistencia.salida_manana_real) if asistencia.salida_manana_real else None,
                "entrada_tarde_real": str(asistencia.entrada_tarde_real) if asistencia.entrada_tarde_real else None,
                "salida_tarde_real": str(asistencia.salida_tarde_real) if asistencia.salida_tarde_real else None
            })
        else:
            return jsonify({}), 200
            
    except Exception as e:
        print(f"Error en api_get_asistencia_empleado: {e}")
        return jsonify({"error": str(e)}), 500

# Asegúrate de tener este import arriba (normalmente ya lo tienes en app.py)
from types import SimpleNamespace 

@app.route('/api/reports/export/excel')
def export_report_excel():
    try:
        empresa_id = request.args.get('empresa_id', type=int)
        mes = request.args.get('mes', type=int, default=datetime.now().month)
        anio = request.args.get('anio', type=int, default=datetime.now().year)
        
        if not empresa_id:
            flash('Empresa ID requerido', 'error')
            return redirect(url_for('reports'))
        
        # 1. Obtener empleados y rango de fechas
        empleados = empleado_repo.get_by_empresa_id(empresa_id)
        _, last_day = calendar.monthrange(anio, mes)
        
        # 2. CONEXIÓN RÁPIDA (BATCH)
        from src.infrastructure.mysql_connection import get_connection
        conn = get_connection()
        cursor = conn.cursor()
        
        # Traemos todo de una vez
        query_batch = """
            SELECT 
                a.empleado_id, 
                DATE(a.fecha) as fecha_dia,
                a.entrada_manana_real, 
                a.salida_manana_real, 
                a.entrada_tarde_real, 
                a.salida_tarde_real
            FROM ASISTENCIA a
            JOIN EMPLEADOS e ON a.empleado_id = e.id
            WHERE e.empresa_id = %s 
              AND YEAR(a.fecha) = %s 
              AND MONTH(a.fecha) = %s
        """
        cursor.execute(query_batch, (empresa_id, anio, mes))
        rows = cursor.fetchall()
        cursor.close()
        conn.close()

        # 3. PROCESAMIENTO Y CACHÉ (AQUÍ ESTÁ LA CORRECCIÓN DEL ERROR)
        asistencia_cache = {}
        
        # Función auxiliar para convertir timedelta (MySQL) a time (Python)
        def convertir_tiempo(valor):
            if isinstance(valor, timedelta):
                # Sumamos el timedelta a la "fecha mínima" para extraer la hora
                return (datetime.min + valor).time()
            return valor

        for row in rows:
            emp_id, fecha_dt, ent_m, sal_m, ent_t, sal_t = row
            key = (emp_id, str(fecha_dt))
            
            # Convertimos los valores antes de guardarlos
            asistencia_obj = SimpleNamespace(
                entrada_manana_real=convertir_tiempo(ent_m),
                salida_manana_real=convertir_tiempo(sal_m),
                entrada_tarde_real=convertir_tiempo(ent_t),
                salida_tarde_real=convertir_tiempo(sal_t)
            )
            asistencia_cache[key] = asistencia_obj

        # 4. GENERACIÓN DEL EXCEL (CÓDIGO ORIGINAL SIN CAMBIOS)
        output = BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte Diario"
        
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        alignment_center = Alignment(horizontal="center", vertical="center")
        
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        fila_actual = 1
        
        for empleado in empleados:
            ws.merge_cells(start_row=fila_actual, start_column=1, end_row=fila_actual, end_column=10)
            titulo_cell = ws.cell(row=fila_actual, column=1, value=f"EMPLEADO: {empleado.nombre.upper()}")
            titulo_cell.font = Font(bold=True, size=14)
            titulo_cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            titulo_cell.alignment = Alignment(horizontal="center", vertical="center")
            fila_actual += 1
            
            headers = ['DIA', 'ENTRADA_MANANA', 'SALIDA_MANANA', 'TOTAL_MANANA', 
                      'ENTRADA_TARDE', 'SALIDA_TARDE', 'TOTAL_TARDE', 'TOTAL_DIA', 'HORAS_NORMALES', 'HORAS_EXTRAS']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=fila_actual, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = alignment_center
                cell.border = thin_border
            fila_actual += 1
            
            total_manana_minutos = 0
            total_tarde_minutos = 0
            total_horas_extras_mes = 0
            
            dias_del_mes = calendar.monthrange(anio, mes)[1]
            
            for dia in range(1, dias_del_mes + 1):
                fecha = f"{anio}-{mes:02d}-{dia:02d}"
                
                # Buscamos en el caché en lugar de la DB
                asistencia = asistencia_cache.get((empleado.id, fecha))
                
                if asistencia:
                    total_manana = ""
                    # Ahora .hour funcionará porque ya convertimos el timedelta a time
                    if asistencia.entrada_manana_real and asistencia.salida_manana_real:
                        minutos_manana = int((asistencia.salida_manana_real.hour * 60 + asistencia.salida_manana_real.minute) - 
                                           (asistencia.entrada_manana_real.hour * 60 + asistencia.entrada_manana_real.minute))
                        total_manana = minutos_a_hhmm(max(0, minutos_manana))
                        h, m = map(int, total_manana.split(":"))
                        total_manana_minutos += h * 60 + m
                    
                    total_tarde = ""
                    if asistencia.entrada_tarde_real and asistencia.salida_tarde_real:
                        minutos_tarde = int((asistencia.salida_tarde_real.hour * 60 + asistencia.salida_tarde_real.minute) - 
                                          (asistencia.entrada_tarde_real.hour * 60 + asistencia.entrada_tarde_real.minute))
                        total_tarde = minutos_a_hhmm(max(0, minutos_tarde))
                        h, m = map(int, total_tarde.split(":"))
                        total_tarde_minutos += h * 60 + m
                    
                    total_dia_minutos = 0
                    if total_manana:
                        h, m = map(int, total_manana.split(":"))
                        total_dia_minutos += h * 60 + m
                    if total_tarde:
                        h, m = map(int, total_tarde.split(":"))
                        total_dia_minutos += h * 60 + m
                    total_dia = minutos_a_hhmm(total_dia_minutos)
                    
                    minutos_normales_dia = min(total_dia_minutos, 8 * 60)
                    minutos_extras_dia = max(0, total_dia_minutos - (8 * 60))
                    
                    total_horas_extras_mes += minutos_extras_dia
                    
                    horas_normales_dia = minutos_a_hhmm(minutos_normales_dia)
                    horas_extras_dia = minutos_a_hhmm(minutos_extras_dia)
                    
                    valores = [
                        dia,
                        str(asistencia.entrada_manana_real) if asistencia.entrada_manana_real else "",
                        str(asistencia.salida_manana_real) if asistencia.salida_manana_real else "",
                        total_manana,
                        str(asistencia.entrada_tarde_real) if asistencia.entrada_tarde_real else "",
                        str(asistencia.salida_tarde_real) if asistencia.salida_tarde_real else "",
                        total_tarde,
                        total_dia,
                        horas_normales_dia,
                        horas_extras_dia
                    ]
                else:
                    valores = [dia, "", "", "", "", "", "", "", "", ""]
                
                for col, valor in enumerate(valores, 1):
                    cell = ws.cell(row=fila_actual, column=col, value=valor)
                    cell.border = thin_border
                    if col == 1:
                        cell.alignment = Alignment(horizontal="center")
                
                fila_actual += 1
            
            # Totales finales (sin cambios)
            total_manana_mes = minutos_a_hhmm(total_manana_minutos)
            total_tarde_mes = minutos_a_hhmm(total_tarde_minutos)
            total_dia_mes = minutos_a_hhmm(total_manana_minutos + total_tarde_minutos)

            total_minutos_mes = total_manana_minutos + total_tarde_minutos
            minutos_normales_mes = total_minutos_mes - total_horas_extras_mes
            minutos_extras_mes = total_horas_extras_mes

            horas_normales_mes = minutos_a_hhmm(minutos_normales_mes)
            horas_extras_mes = minutos_a_hhmm(minutos_extras_mes)
            
            valores_totales = [
                "TOTAL MES", "", "", total_manana_mes, "", "", total_tarde_mes,
                total_dia_mes, horas_normales_mes, horas_extras_mes
            ]
            for col, valor in enumerate(valores_totales, 1):
                cell = ws.cell(row=fila_actual, column=col, value=valor)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
                cell.border = thin_border
                if col == 1:
                    cell.alignment = Alignment(horizontal="center")
            fila_actual += 2
        
        # Ajuste de ancho de columnas
        column_widths = [8, 15, 15, 15, 15, 15, 15, 15, 15, 15]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
        
        wb.save(output)
        output.seek(0)
        
        nombre_empresa = empresa_repo.get_by_id(empresa_id).nombre.replace(' ', '_') if empresa_repo.get_by_id(empresa_id) else 'empresa'
        nombre_archivo = f"reporte_diario_{nombre_empresa}_{mes}_{anio}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=nombre_archivo
        )
        
    except Exception as e:
        # Esto imprimirá el error real en los logs de Fly.io
        print(f"ERROR CRÍTICO EXPORT EXCEL: {str(e)}")
        import traceback
        print(traceback.format_exc())
        flash(f'Error generando reporte Excel: {str(e)}', 'error')
        return redirect(url_for('reports'))

@app.route('/admin/generate_qr/<int:empleado_id>')
def generate_employee_qr(empleado_id):
    if not session.get('admin_logged_in'):
        return redirect(url_for('admin_login'))
    
    try:
        empleado = empleado_repo.get_by_id(empleado_id)
        if not empleado:
            flash('Empleado no encontrado', 'error')
            return redirect(url_for('admin_list_employees'))
        
        empresa = empresa_repo.get_by_id(empleado.empresa_id)
        if not empresa:
            flash('Empresa no encontrada', 'error')
            return redirect(url_for('admin_list_employees'))
        
        qr_path = qr_generator.generate_employee_qr(empleado.id, empresa.codigo_empresa)
        if qr_path:
            flash('Código QR generado con éxito.', 'success')
        else:
            flash('Error generando código QR', 'error')
            
    except Exception as e:
        flash(f'Error: {str(e)}', 'error')
    
    return redirect(url_for('admin_list_employees'))

@app.errorhandler(404)
def not_found(error):
    return render_template('error.html', error_message="Página no encontrada"), 404

@app.errorhandler(500)
def internal_error(error):
    return render_template('error.html', error_message="Error interno del servidor"), 500

# ===============================================
# 📊 REPORTE SEMANAL DE ACTIVIDAD Y RENDIMIENTO
# ===============================================

from datetime import timedelta
from collections import Counter

@app.route('/admin/weekly-report')
def admin_weekly_report():
    """Página principal del reporte semanal"""
    if not session.get('admin_logged_in'):
        flash('Debes iniciar sesión para acceder al reporte semanal', 'error')
        return redirect(url_for('admin_login'))
    
    empresas = list_companies_use_case.execute()
    return render_template('weekly_report.html', empresas=empresas)

@app.route('/api/weekly-report/daily-attendance')
def api_weekly_report_daily_attendance():
    """Asistencia diaria - MODIFICADO"""
    if not session.get('admin_logged_in'):
        return jsonify({"error": "No autorizado"}), 401
    
    try:
        from src.infrastructure.mysql_connection import get_connection
        
        empresa_id = request.args.get('empresa_id', type=int)
        fecha_inicio = request.args.get('fecha_inicio')
        fecha_fin = request.args.get('fecha_fin')
        
        if not fecha_inicio or not fecha_fin:
            return jsonify({"error": "Fechas requeridas"}), 400
        
        conn = get_connection()
        cursor = conn.cursor()
        
        dias_labels = []
        asistencias_data = []
        tardanzas_data = []
        faltas_data = []
        
        # Total empleados
        if empresa_id:
            cursor.execute("SELECT COUNT(*) FROM EMPLEADOS WHERE empresa_id = %s AND activo = TRUE", (empresa_id,))
        else:
            cursor.execute("SELECT COUNT(*) FROM EMPLEADOS WHERE activo = TRUE")
        total_empleados = cursor.fetchone()[0]
        
        # Iterar días
        fecha_actual = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
        fecha_final = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
        
        while fecha_actual <= fecha_final:
            if fecha_actual > datetime.now().date():
                break
            
            dia_str = fecha_actual.strftime('%Y-%m-%d')
            
            dias_semana_es = {
                'Monday': 'Lun', 'Tuesday': 'Mar', 'Wednesday': 'Mié',
                'Thursday': 'Jue', 'Friday': 'Vie', 'Saturday': 'Sáb', 'Sunday': 'Dom'
            }
            dia_nombre = dias_semana_es.get(fecha_actual.strftime('%A'), fecha_actual.strftime('%A')[:3])
            dias_labels.append(f"{dia_nombre} {fecha_actual.day}")
            
            empresa_filter = "AND e.empresa_id = %s" if empresa_id else ""
            params = [dia_str]
            if empresa_id:
                params.append(empresa_id)
            
            # Asistencias (empleados únicos)
            cursor.execute(f"""
                SELECT COUNT(DISTINCT a.empleado_id)
                FROM ASISTENCIA a
                JOIN EMPLEADOS e ON a.empleado_id = e.id
                WHERE a.fecha = %s
                  AND (a.entrada_manana_real IS NOT NULL OR a.entrada_tarde_real IS NOT NULL)
                  {empresa_filter}
            """, params)
            asistencias = cursor.fetchone()[0]
            asistencias_data.append(asistencias)
            
            # 🔥 Tardanzas (contar turnos con tardanza)
            cursor.execute(f"""
                SELECT 
                    (COUNT(CASE WHEN a.entrada_manana_real IS NOT NULL 
                                AND TIME(a.entrada_manana_real) > '06:50:59' THEN 1 END) +
                     COUNT(CASE WHEN a.entrada_tarde_real IS NOT NULL 
                                AND TIME(a.entrada_tarde_real) > '14:50:59' THEN 1 END))
                FROM ASISTENCIA a
                JOIN EMPLEADOS e ON a.empleado_id = e.id
                WHERE a.fecha = %s
                  {empresa_filter}
            """, params)
            tardanzas = cursor.fetchone()[0] or 0
            tardanzas_data.append(tardanzas)
            
            faltas = total_empleados - asistencias
            faltas_data.append(faltas)
            
            fecha_actual += timedelta(days=1)
        
        cursor.close()
        conn.close()
        
        return jsonify({
            "dias": dias_labels,
            "asistencias": asistencias_data,
            "tardanzas": tardanzas_data,
            "faltas": faltas_data
        })
        
    except Exception as e:
        import traceback
        print(f"❌ Error en daily-attendance: {e}")
        print(traceback.format_exc())
        return jsonify({"error": str(e)}), 500
#HOLIIIIIIIII
@app.route('/api/weekly-report/daily-attendance-details')
def api_weekly_report_daily_attendance_details():
    """Detalles de asistencia diaria CON NOMBRES para tooltips"""
    if not session.get('admin_logged_in'):
        return jsonify({"error": "No autorizado"}), 401

    try:
        from src.infrastructure.mysql_connection import get_connection

        empresa_id = request.args.get('empresa_id', type=int)
        fecha_inicio = request.args.get('fecha_inicio')
        fecha_fin = request.args.get('fecha_fin')

        if not fecha_inicio or not fecha_fin:
            return jsonify({"error": "Fechas requeridas"}), 400

        conn = get_connection()
        cursor = conn.cursor()

        empresa_filter = "AND e.empresa_id = %s" if empresa_id else ""

        # Total empleados activos
        if empresa_id:
            cursor.execute("SELECT id, nombre FROM EMPLEADOS WHERE empresa_id = %s AND activo = TRUE", (empresa_id,))
        else:
            cursor.execute("SELECT id, nombre FROM EMPLEADOS WHERE activo = TRUE")

        todos_empleados = {row[0]: row[1] for row in cursor.fetchall()}

        resultado = {}

        # Iterar días
        fecha_actual = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
        fecha_final = datetime.strptime(fecha_fin, '%Y-%m-%d').date()

        while fecha_actual <= fecha_final:
            if fecha_actual > datetime.now().date():
                break

            dia_str = fecha_actual.strftime('%Y-%m-%d')

            params = [dia_str]
            if empresa_id:
                params.append(empresa_id)

            # 🔥 Obtener TODOS los registros del día con nombres
            cursor.execute(f"""
                SELECT 
                    e.id,
                    e.nombre,
                    TIME(a.entrada_manana_real) as entrada_manana,
                    TIME(a.entrada_tarde_real) as entrada_tarde
                FROM EMPLEADOS e
                LEFT JOIN ASISTENCIA a ON e.id = a.empleado_id AND a.fecha = %s
                WHERE e.activo = TRUE {empresa_filter}
                ORDER BY e.nombre
            """, params)

            registros = cursor.fetchall()

            puntuales = []
            tardes_manana = []
            tardes_tarde = []
            faltas = []

            for empleado_id, nombre, entrada_manana, entrada_tarde in registros:
                # Si NO tiene ningún registro
                if entrada_manana is None and entrada_tarde is None:
                    faltas.append(nombre)
                    continue

                # 🕕 Verificar mañana
                if entrada_manana:
                    # Convertir timedelta a time si es necesario
                    if isinstance(entrada_manana, timedelta):
                        entrada_manana = (datetime.min + entrada_manana).time()

                    hora_manana = entrada_manana.strftime('%H:%M:%S')
                    if entrada_manana <= datetime.strptime('06:50:59', '%H:%M:%S').time():
                        puntuales.append(f"{nombre} (M)")
                    else:
                        tardes_manana.append(f"{nombre} ({hora_manana})")

                # 🌇 Verificar tarde
                if entrada_tarde:
                    if isinstance(entrada_tarde, timedelta):
                        entrada_tarde = (datetime.min + entrada_tarde).time()

                    hora_tarde = entrada_tarde.strftime('%H:%M:%S')
                    if entrada_tarde <= datetime.strptime('14:50:59', '%H:%M:%S').time():
                        if f"{nombre} (M)" not in puntuales:
                            puntuales.append(f"{nombre} (T)")
                    else:
                        tardes_tarde.append(f"{nombre} ({hora_tarde})")

            # 🗓️ Etiqueta del día (Lun, Mar, Mié, etc.)
            dias_semana_es = {
                'Monday': 'Lun', 'Tuesday': 'Mar', 'Wednesday': 'Mié',
                'Thursday': 'Jue', 'Friday': 'Vie', 'Saturday': 'Sáb', 'Sunday': 'Dom'
            }
            dia_nombre = dias_semana_es.get(fecha_actual.strftime('%A'), fecha_actual.strftime('%A')[:3])
            label = f"{dia_nombre} {fecha_actual.day}"

            resultado[label] = {
                "puntuales": puntuales,
                "tardes_manana": tardes_manana,
                "tardes_tarde": tardes_tarde,
                "faltas": faltas,
                "total_asistencias": len(todos_empleados) - len(faltas),
                "total_tardanzas": len(tardes_manana) + len(tardes_tarde),
                "total_faltas": len(faltas)
            }

            fecha_actual += timedelta(days=1)

        cursor.close()
        conn.close()

        return jsonify(resultado)

    except Exception as e:
        import traceback
        print(f"❌ Error en daily-attendance-details: {e}")
        print(traceback.format_exc())
        return jsonify({"error": str(e)}), 500



# ========================================
# PASO 4: MODIFICAR /api/weekly-report/frequent-hours
# (Solo si NO acepta fecha_inicio/fecha_fin)
# ========================================

@app.route('/api/weekly-report/frequent-hours')
def api_weekly_report_frequent_hours():
    """Horas frecuentes - MEJORADO: hora exacta más común"""
    if not session.get('admin_logged_in'):
        return jsonify({"error": "No autorizado"}), 401
    
    try:
        from src.infrastructure.mysql_connection import get_connection
        
        empresa_id = request.args.get('empresa_id', type=int)
        fecha_inicio = request.args.get('fecha_inicio')
        fecha_fin = request.args.get('fecha_fin')
        
        if not fecha_inicio or not fecha_fin:
            return jsonify({"error": "Fechas requeridas"}), 400
        
        conn = get_connection()
        cursor = conn.cursor()
        
        empresa_filter = "AND e.empresa_id = %s" if empresa_id else ""
        params = [fecha_inicio, fecha_fin]
        if empresa_id:
            params.append(empresa_id)
        
        # 🔥 Hora mañana exacta (HH:MM)
        cursor.execute(f"""
            SELECT TIME_FORMAT(a.entrada_manana_real, '%H:%i') as hora, COUNT(*) as freq
            FROM ASISTENCIA a
            JOIN EMPLEADOS e ON a.empleado_id = e.id
            WHERE a.fecha BETWEEN %s AND %s
              AND a.entrada_manana_real IS NOT NULL
              {empresa_filter}
            GROUP BY hora
            ORDER BY freq DESC
            LIMIT 1
        """, params)
        
        result_manana = cursor.fetchone()
        hora_frecuente_manana = result_manana[0] if result_manana else "N/A"
        frecuencia_manana = result_manana[1] if result_manana else 0
        
        # 🔥 Hora tarde exacta (HH:MM)
        cursor.execute(f"""
            SELECT TIME_FORMAT(a.entrada_tarde_real, '%H:%i') as hora, COUNT(*) as freq
            FROM ASISTENCIA a
            JOIN EMPLEADOS e ON a.empleado_id = e.id
            WHERE a.fecha BETWEEN %s AND %s
              AND a.entrada_tarde_real IS NOT NULL
              {empresa_filter}
            GROUP BY hora
            ORDER BY freq DESC
            LIMIT 1
        """, params)
        
        result_tarde = cursor.fetchone()
        hora_frecuente_tarde = result_tarde[0] if result_tarde else "N/A"
        frecuencia_tarde = result_tarde[1] if result_tarde else 0
        
        cursor.close()
        conn.close()
        
        return jsonify({
            "hora_frecuente_manana": hora_frecuente_manana,
            "frecuencia_manana": frecuencia_manana,
            "hora_frecuente_tarde": hora_frecuente_tarde,
            "frecuencia_tarde": frecuencia_tarde
        })
        
    except Exception as e:
        import traceback
        print(f"❌ Error en frequent-hours: {e}")
        print(traceback.format_exc())
        return jsonify({"error": str(e)}), 500
    
@app.route('/api/weekly-report/top-punctual-morning')
def api_weekly_report_top_punctual_morning():
    """Top puntuales turno mañana - BASADO EN HORA REAL"""
    if not session.get('admin_logged_in'):
        return jsonify({"error": "No autorizado"}), 401
    
    try:
        from src.infrastructure.mysql_connection import get_connection
        
        empresa_id = request.args.get('empresa_id', type=int)
        fecha_inicio = request.args.get('fecha_inicio')
        fecha_fin = request.args.get('fecha_fin')
        
        if not fecha_inicio or not fecha_fin:
            return jsonify({"error": "Fechas requeridas"}), 400
        
        conn = get_connection()
        cursor = conn.cursor()
        
        empresa_filter = "AND e.empresa_id = %s" if empresa_id else ""
        params = [fecha_inicio, fecha_fin]
        if empresa_id:
            params.append(empresa_id)
        
        # Hora real <= 06:50:59
        cursor.execute(f"""
            SELECT 
                e.nombre,
                COUNT(CASE 
                    WHEN a.entrada_manana_real IS NOT NULL 
                         AND TIME(a.entrada_manana_real) <= '06:50:59'
                    THEN 1 
                END) as puntualidades,
                COUNT(CASE 
                    WHEN a.entrada_manana_real IS NOT NULL 
                    THEN 1 
                END) as total_ingresos
            FROM EMPLEADOS e
            LEFT JOIN ASISTENCIA a ON e.id = a.empleado_id 
                AND a.fecha BETWEEN %s AND %s
            WHERE e.activo = TRUE {empresa_filter}
            GROUP BY e.id, e.nombre
            HAVING puntualidades > 0
            ORDER BY puntualidades DESC, total_ingresos DESC
            LIMIT 5
        """, params)
        
        result = []
        for row in cursor.fetchall():
            result.append({
                "nombre": row[0],
                "puntualidades": row[1],
                "total_ingresos": row[2]
            })
        
        cursor.close()
        conn.close()
        return jsonify(result)
        
    except Exception as e:
        import traceback
        print(f"❌ Error en top-punctual-morning: {e}")
        print(traceback.format_exc())
        return jsonify({"error": str(e)}), 500


@app.route('/api/weekly-report/top-punctual-afternoon')
def api_weekly_report_top_punctual_afternoon():
    """Top puntuales turno tarde - BASADO EN HORA REAL"""
    if not session.get('admin_logged_in'):
        return jsonify({"error": "No autorizado"}), 401
    
    try:
        from src.infrastructure.mysql_connection import get_connection
        
        empresa_id = request.args.get('empresa_id', type=int)
        fecha_inicio = request.args.get('fecha_inicio')
        fecha_fin = request.args.get('fecha_fin')
        
        if not fecha_inicio or not fecha_fin:
            return jsonify({"error": "Fechas requeridas"}), 400
        
        conn = get_connection()
        cursor = conn.cursor()
        
        empresa_filter = "AND e.empresa_id = %s" if empresa_id else ""
        params = [fecha_inicio, fecha_fin]
        if empresa_id:
            params.append(empresa_id)
        
        #  Hora real <= 14:50:59
        cursor.execute(f"""
            SELECT 
                e.nombre,
                COUNT(CASE 
                    WHEN a.entrada_tarde_real IS NOT NULL 
                         AND TIME(a.entrada_tarde_real) <= '14:50:59'
                    THEN 1 
                END) as puntualidades,
                COUNT(CASE 
                    WHEN a.entrada_tarde_real IS NOT NULL 
                    THEN 1 
                END) as total_ingresos
            FROM EMPLEADOS e
            LEFT JOIN ASISTENCIA a ON e.id = a.empleado_id 
                AND a.fecha BETWEEN %s AND %s
            WHERE e.activo = TRUE {empresa_filter}
            GROUP BY e.id, e.nombre
            HAVING puntualidades > 0
            ORDER BY puntualidades DESC, total_ingresos DESC
            LIMIT 5
        """, params)
        
        result = []
        for row in cursor.fetchall():
            result.append({
                "nombre": row[0],
                "puntualidades": row[1],
                "total_ingresos": row[2]
            })
        
        cursor.close()
        conn.close()
        return jsonify(result)
        
    except Exception as e:
        import traceback
        print(f"❌ Error en top-punctual-afternoon: {e}")
        print(traceback.format_exc())
        return jsonify({"error": str(e)}), 500


@app.route('/api/weekly-report/top-late-morning')
def api_weekly_report_top_late_morning():
    """Top tardones turno mañana - BASADO EN HORA REAL"""
    if not session.get('admin_logged_in'):
        return jsonify({"error": "No autorizado"}), 401
    
    try:
        from src.infrastructure.mysql_connection import get_connection
        
        empresa_id = request.args.get('empresa_id', type=int)
        fecha_inicio = request.args.get('fecha_inicio')
        fecha_fin = request.args.get('fecha_fin')
        
        if not fecha_inicio or not fecha_fin:
            return jsonify({"error": "Fechas requeridas"}), 400
        
        conn = get_connection()
        cursor = conn.cursor()
        
        empresa_filter = "AND e.empresa_id = %s" if empresa_id else ""
        params = [fecha_inicio, fecha_fin]
        if empresa_id:
            params.append(empresa_id)
        
        #  Hora real > 06:50:59
        cursor.execute(f"""
            SELECT 
                e.nombre,
                COUNT(CASE 
                    WHEN a.entrada_manana_real IS NOT NULL 
                         AND TIME(a.entrada_manana_real) > '06:50:59'
                    THEN 1 
                END) as tardanzas,
                COUNT(CASE 
                    WHEN a.entrada_manana_real IS NOT NULL 
                    THEN 1 
                END) as total_ingresos
            FROM EMPLEADOS e
            INNER JOIN ASISTENCIA a ON e.id = a.empleado_id 
                AND a.fecha BETWEEN %s AND %s
            WHERE e.activo = TRUE {empresa_filter}
            GROUP BY e.id, e.nombre
            HAVING tardanzas > 0
            ORDER BY tardanzas DESC, total_ingresos DESC
            LIMIT 5
        """, params)
        
        result = []
        for row in cursor.fetchall():
            result.append({
                "nombre": row[0],
                "tardanzas": row[1],
                "total_ingresos": row[2]
            })
        
        cursor.close()
        conn.close()
        return jsonify(result)
        
    except Exception as e:
        import traceback
        print(f"❌ Error en top-late-morning: {e}")
        print(traceback.format_exc())
        return jsonify({"error": str(e)}), 500


@app.route('/api/weekly-report/top-late-afternoon')
def api_weekly_report_top_late_afternoon():
    """Top tardones turno tarde - BASADO EN HORA REAL"""
    if not session.get('admin_logged_in'):
        return jsonify({"error": "No autorizado"}), 401
    
    try:
        from src.infrastructure.mysql_connection import get_connection
        
        empresa_id = request.args.get('empresa_id', type=int)
        fecha_inicio = request.args.get('fecha_inicio')
        fecha_fin = request.args.get('fecha_fin')
        
        if not fecha_inicio or not fecha_fin:
            return jsonify({"error": "Fechas requeridas"}), 400
        
        conn = get_connection()
        cursor = conn.cursor()
        
        empresa_filter = "AND e.empresa_id = %s" if empresa_id else ""
        params = [fecha_inicio, fecha_fin]
        if empresa_id:
            params.append(empresa_id)
        
        # 🔥 Hora real > 14:50:59
        cursor.execute(f"""
            SELECT 
                e.nombre,
                COUNT(CASE 
                    WHEN a.entrada_tarde_real IS NOT NULL 
                         AND TIME(a.entrada_tarde_real) > '14:50:59'
                    THEN 1 
                END) as tardanzas,
                COUNT(CASE 
                    WHEN a.entrada_tarde_real IS NOT NULL 
                    THEN 1 
                END) as total_ingresos
            FROM EMPLEADOS e
            INNER JOIN ASISTENCIA a ON e.id = a.empleado_id 
                AND a.fecha BETWEEN %s AND %s
            WHERE e.activo = TRUE {empresa_filter}
            GROUP BY e.id, e.nombre
            HAVING tardanzas > 0
            ORDER BY tardanzas DESC, total_ingresos DESC
            LIMIT 5
        """, params)
        
        result = []
        for row in cursor.fetchall():
            result.append({
                "nombre": row[0],
                "tardanzas": row[1],
                "total_ingresos": row[2]
            })
        
        cursor.close()
        conn.close()
        return jsonify(result)
        
    except Exception as e:
        import traceback
        print(f"❌ Error en top-late-afternoon: {e}")
        print(traceback.format_exc())
        return jsonify({"error": str(e)}), 500


# ========================================
# PASO 2: MODIFICAR el endpoint /api/weekly-report/summary
# (Buscar este endpoint en tu app.py y REEMPLAZARLO)
# ========================================

@app.route('/api/weekly-report/summary')
def api_weekly_report_summary():
    """Resumen general - VERSIÓN MEJORADA con desglose de tardanzas"""
    if not session.get('admin_logged_in'):
        return jsonify({"error": "No autorizado"}), 401
    
    try:
        from src.infrastructure.mysql_connection import get_connection
        
        empresa_id = request.args.get('empresa_id', type=int)
        fecha_inicio = request.args.get('fecha_inicio')
        fecha_fin = request.args.get('fecha_fin')
        
        if not fecha_inicio or not fecha_fin:
            semana_offset = request.args.get('semana', type=int, default=0)
            hoy = datetime.now().date()
            inicio_semana = hoy - timedelta(days=hoy.weekday()) + timedelta(weeks=semana_offset)
            fin_semana = inicio_semana + timedelta(days=6)
            fecha_inicio = inicio_semana.strftime('%Y-%m-%d')
            fecha_fin = fin_semana.strftime('%Y-%m-%d')
        
        conn = get_connection()
        cursor = conn.cursor()
        
        empresa_filter = "AND e.empresa_id = %s" if empresa_id else ""
        params = [fecha_inicio, fecha_fin]
        if empresa_id:
            params.append(empresa_id)
        
        # Total empleados activos
        if empresa_id:
            cursor.execute("SELECT COUNT(*) FROM EMPLEADOS WHERE empresa_id = %s AND activo = TRUE", (empresa_id,))
        else:
            cursor.execute("SELECT COUNT(*) FROM EMPLEADOS WHERE activo = TRUE")
        total_empleados = cursor.fetchone()[0]
        
        # Contar turnos para puntualidad y asistencia
        cursor.execute(f"""
            SELECT 
                (COUNT(CASE WHEN a.entrada_manana_real IS NOT NULL THEN 1 END) + 
                 COUNT(CASE WHEN a.entrada_tarde_real IS NOT NULL THEN 1 END)) as registros_totales,
                
                (COUNT(CASE WHEN a.entrada_manana_real IS NOT NULL 
                            AND TIME(a.entrada_manana_real) <= '06:50:59' THEN 1 END) +
                 COUNT(CASE WHEN a.entrada_tarde_real IS NOT NULL 
                            AND TIME(a.entrada_tarde_real) <= '14:50:59' THEN 1 END)) as registros_puntuales
            FROM ASISTENCIA a
            JOIN EMPLEADOS e ON a.empleado_id = e.id
            WHERE a.fecha BETWEEN %s AND %s
              AND e.activo = TRUE
              {empresa_filter}
        """, params)
        
        result = cursor.fetchone()
        registros_totales = result[0] or 0
        registros_puntuales = result[1] or 0
        
        promedio_puntualidad = int((registros_puntuales / registros_totales * 100)) if registros_totales > 0 else 0
        
        # Calcular días del período
        fecha_inicio_dt = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
        fecha_fin_dt = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
        dias_periodo = (fecha_fin_dt - fecha_inicio_dt).days + 1
        dias_transcurridos = min(dias_periodo, (datetime.now().date() - fecha_inicio_dt).days + 1)
        if dias_transcurridos < 1:
            dias_transcurridos = 1
        
        # Turnos esperados
        turnos_esperados = total_empleados * dias_transcurridos * 2
        porcentaje_asistencia = int((registros_totales / turnos_esperados * 100)) if turnos_esperados > 0 else 0
        
        # 🔥 TARDANZAS CON DESGLOSE POR TURNO
        cursor.execute(f"""
            SELECT 
                COUNT(CASE 
                    WHEN a.entrada_manana_real IS NOT NULL 
                         AND TIME(a.entrada_manana_real) > '06:50:59'
                    THEN 1 
                END) as tardanzas_manana,
                COUNT(CASE 
                    WHEN a.entrada_tarde_real IS NOT NULL 
                         AND TIME(a.entrada_tarde_real) > '14:50:59'
                    THEN 1 
                END) as tardanzas_tarde
            FROM ASISTENCIA a
            JOIN EMPLEADOS e ON a.empleado_id = e.id
            WHERE a.fecha BETWEEN %s AND %s
              AND e.activo = TRUE
              {empresa_filter}
        """, params)
        
        result_tardanzas = cursor.fetchone()
        tardanzas_manana = result_tardanzas[0] or 0
        tardanzas_tarde = result_tardanzas[1] or 0
        total_tardanzas = tardanzas_manana + tardanzas_tarde
        
        # Total faltas
        total_faltas = turnos_esperados - registros_totales
        
        # Horas extras
        cursor.execute(f"""
            SELECT SUM(a.horas_extras)
            FROM ASISTENCIA a
            JOIN EMPLEADOS e ON a.empleado_id = e.id
            WHERE a.fecha BETWEEN %s AND %s
              {empresa_filter}
        """, params)
        horas_extras = cursor.fetchone()[0] or 0
        
        cursor.close()
        conn.close()
        
        return jsonify({
            "periodo": {
                "inicio": fecha_inicio,
                "fin": fecha_fin,
                "inicio_formato": datetime.strptime(fecha_inicio, '%Y-%m-%d').strftime('%d/%m/%Y'),
                "fin_formato": datetime.strptime(fecha_fin, '%Y-%m-%d').strftime('%d/%m/%Y')
            },
            "total_empleados": total_empleados,
            "promedio_puntualidad": promedio_puntualidad,
            "porcentaje_asistencia": porcentaje_asistencia,
            "total_tardanzas": total_tardanzas,  # 🎯 Total general
            "tardanzas_manana": tardanzas_manana,  # 📊 Desglose
            "tardanzas_tarde": tardanzas_tarde,     # 📊 Desglose
            "total_faltas": total_faltas,
            "horas_extras": round(horas_extras, 2),
            "dias_periodo": dias_periodo
        })
        
    except Exception as e:
        import traceback
        print(f"❌ Error en summary: {e}")
        print(traceback.format_exc())
        return jsonify({"error": str(e)}), 500


@app.route('/api/weekly-report/worst-days')
def api_weekly_report_worst_days():
    """Días con menor asistencia"""
    if not session.get('admin_logged_in'):
        return jsonify({"error": "No autorizado"}), 401
    
    try:
        from src.infrastructure.mysql_connection import get_connection
        
        empresa_id = request.args.get('empresa_id', type=int)
        semana_offset = request.args.get('semana', type=int, default=0)
        
        hoy = datetime.now().date()
        inicio_semana = hoy - timedelta(days=hoy.weekday()) + timedelta(weeks=semana_offset)
        fin_semana = inicio_semana + timedelta(days=6)
        
        conn = get_connection()
        cursor = conn.cursor()
        
        empresa_filter = "AND e.empresa_id = %s" if empresa_id else ""
        params = [inicio_semana.strftime('%Y-%m-%d'), fin_semana.strftime('%Y-%m-%d')]
        if empresa_id:
            params.append(empresa_id)
        
        cursor.execute(f"""
            SELECT 
                a.fecha,
                COUNT(DISTINCT a.empleado_id) as total_asistencias,
                DAYNAME(a.fecha) as dia_semana
            FROM ASISTENCIA a
            JOIN EMPLEADOS e ON a.empleado_id = e.id
            WHERE a.fecha BETWEEN %s AND %s
              AND (a.entrada_manana_real IS NOT NULL OR a.entrada_tarde_real IS NOT NULL)
              {empresa_filter}
            GROUP BY a.fecha
            ORDER BY total_asistencias ASC
            LIMIT 3
        """, params)
        
        peores_dias = []
        for row in cursor.fetchall():
            fecha = row[0]
            asistencias = row[1]
            dia_semana = row[2]
            
            dias_traduccion = {
                'Monday': 'Lunes', 'Tuesday': 'Martes', 'Wednesday': 'Miércoles',
                'Thursday': 'Jueves', 'Friday': 'Viernes', 'Saturday': 'Sábado', 'Sunday': 'Domingo'
            }
            dia_es = dias_traduccion.get(dia_semana, dia_semana)
            
            peores_dias.append({
                "fecha": fecha.strftime('%d/%m/%Y'),
                "dia": dia_es,
                "asistencias": asistencias
            })
        
        cursor.close()
        conn.close()
        
        return jsonify(peores_dias)
        
    except Exception as e:
        import traceback
        print(f"Error en worst days: {e}")
        print(traceback.format_exc())
        return jsonify({"error": str(e)}), 500


@app.route('/api/weekly-report/companies-comparison')
def api_weekly_report_companies_comparison():
    """Comparación entre empresas"""
    if not session.get('admin_logged_in'):
        return jsonify({"error": "No autorizado"}), 401
    
    try:
        from src.infrastructure.mysql_connection import get_connection
        
        empresa_id = request.args.get('empresa_id', type=int)
        semana_offset = request.args.get('semana', type=int, default=0)
        
        hoy = datetime.now().date()
        inicio_semana = hoy - timedelta(days=hoy.weekday()) + timedelta(weeks=semana_offset)
        fin_semana = inicio_semana + timedelta(days=6)
        
        conn = get_connection()
        cursor = conn.cursor()
        
        # Si hay empresa seleccionada, solo mostrar esa
        if empresa_id:
            cursor.execute("""
                SELECT 
                    emp.id,
                    emp.nombre,
                    COUNT(DISTINCT e.id) as total_empleados,
                    COUNT(DISTINCT CASE WHEN a.fecha BETWEEN %s AND %s 
                                        AND (a.entrada_manana_real IS NOT NULL OR a.entrada_tarde_real IS NOT NULL)
                                        THEN CONCAT(a.empleado_id, '-', a.fecha) END) as asistencias_semana
                FROM EMPRESAS emp
                LEFT JOIN EMPLEADOS e ON emp.id = e.empresa_id AND e.activo = TRUE
                LEFT JOIN ASISTENCIA a ON e.id = a.empleado_id
                WHERE emp.id = %s
                GROUP BY emp.id, emp.nombre
            """, (inicio_semana.strftime('%Y-%m-%d'), fin_semana.strftime('%Y-%m-%d'), empresa_id))
        else:
            cursor.execute("""
                SELECT 
                    emp.id,
                    emp.nombre,
                    COUNT(DISTINCT e.id) as total_empleados,
                    COUNT(DISTINCT CASE WHEN a.fecha BETWEEN %s AND %s 
                                        AND (a.entrada_manana_real IS NOT NULL OR a.entrada_tarde_real IS NOT NULL)
                                        THEN CONCAT(a.empleado_id, '-', a.fecha) END) as asistencias_semana
                FROM EMPRESAS emp
                LEFT JOIN EMPLEADOS e ON emp.id = e.empresa_id AND e.activo = TRUE
                LEFT JOIN ASISTENCIA a ON e.id = a.empleado_id
                GROUP BY emp.id, emp.nombre
                ORDER BY emp.nombre
            """, (inicio_semana.strftime('%Y-%m-%d'), fin_semana.strftime('%Y-%m-%d')))
        
        empresas_data = []
        dias_transcurridos = min(7, (datetime.now().date() - inicio_semana).days + 1)
        if dias_transcurridos < 1:
            dias_transcurridos = 1
        
        for row in cursor.fetchall():
            nombre = row[1]
            total_empleados = row[2] or 0
            asistencias = row[3] or 0
            
            asistencias_esperadas = total_empleados * dias_transcurridos
            porcentaje = int((asistencias / asistencias_esperadas * 100)) if asistencias_esperadas > 0 else 0
            
            if total_empleados > 0:
                empresas_data.append({
                    "nombre": nombre,
                    "total_empleados": total_empleados,
                    "porcentaje_asistencia": porcentaje
                })
        
        cursor.close()
        conn.close()
        
        return jsonify(empresas_data)
        
    except Exception as e:
        import traceback
        print(f"Error en companies comparison: {e}")
        print(traceback.format_exc())
        return jsonify({"error": str(e)}), 500


@app.route('/api/weekly-report/top-punctual')
def api_weekly_report_top_punctual():
    """Top 5 MÁS PUNTUALES - Solo 100% puntuales (0 tardanzas)"""
    if not session.get('admin_logged_in'):
        return jsonify({"error": "No autorizado"}), 401
    
    try:
        from src.infrastructure.mysql_connection import get_connection
        
        empresa_id = request.args.get('empresa_id', type=int)
        fecha_inicio = request.args.get('fecha_inicio')
        fecha_fin = request.args.get('fecha_fin')
        
        if not fecha_inicio or not fecha_fin:
            semana_offset = request.args.get('semana', type=int, default=0)
            hoy = datetime.now().date()
            inicio_semana = hoy - timedelta(days=hoy.weekday()) + timedelta(weeks=semana_offset)
            fin_semana = inicio_semana + timedelta(days=6)
            fecha_inicio = inicio_semana.strftime('%Y-%m-%d')
            fecha_fin = fin_semana.strftime('%Y-%m-%d')
        
        conn = get_connection()
        cursor = conn.cursor()
        
        empresa_filter = "AND e.empresa_id = %s" if empresa_id else ""
        params = [fecha_inicio, fecha_fin]
        if empresa_id:
            params.append(empresa_id)
        
        # 🔥 SOLO empleados con 0 tardanzas
        cursor.execute(f"""
            SELECT 
                e.nombre,
                (
                    COUNT(CASE 
                        WHEN a.entrada_manana_real IS NOT NULL 
                             AND TIME(a.entrada_manana_real) <= '06:50:59'
                        THEN 1 
                    END) +
                    COUNT(CASE 
                        WHEN a.entrada_tarde_real IS NOT NULL 
                             AND TIME(a.entrada_tarde_real) <= '14:50:59'
                        THEN 1 
                    END)
                ) as turnos_puntuales,
                (
                    COUNT(CASE WHEN a.entrada_manana_real IS NOT NULL THEN 1 END) +
                    COUNT(CASE WHEN a.entrada_tarde_real IS NOT NULL THEN 1 END)
                ) as total_turnos,
                (
                    COUNT(CASE 
                        WHEN a.entrada_manana_real IS NOT NULL 
                             AND TIME(a.entrada_manana_real) > '06:50:59'
                        THEN 1 
                    END) +
                    COUNT(CASE 
                        WHEN a.entrada_tarde_real IS NOT NULL 
                             AND TIME(a.entrada_tarde_real) > '14:50:59'
                        THEN 1 
                    END)
                ) as tardanzas
            FROM EMPLEADOS e
            LEFT JOIN ASISTENCIA a ON e.id = a.empleado_id 
                AND a.fecha BETWEEN %s AND %s
            WHERE e.activo = TRUE {empresa_filter}
            GROUP BY e.id, e.nombre
            HAVING turnos_puntuales > 0 
                AND tardanzas = 0 
                AND total_turnos = turnos_puntuales  -- 🎯 100% puntual
            ORDER BY turnos_puntuales DESC, total_turnos DESC
            LIMIT 5
        """, params)
        
        top_puntuales = []
        for row in cursor.fetchall():
            top_puntuales.append({
                "nombre": row[0],
                "turnos_puntuales": row[1],
                "total_turnos": row[2],
                "perfecto": True
            })
        
        cursor.close()
        conn.close()
        
        return jsonify(top_puntuales)
        
    except Exception as e:
        import traceback
        print(f"❌ Error en top punctual: {e}")
        print(traceback.format_exc())
        return jsonify({"error": str(e)}), 500


@app.route('/api/weekly-report/top-late')
def api_weekly_report_top_late():
    """Top 5 MÁS TARDONES - Empleados con al menos 1 tardanza"""
    if not session.get('admin_logged_in'):
        return jsonify({"error": "No autorizado"}), 401
    
    try:
        from src.infrastructure.mysql_connection import get_connection
        
        empresa_id = request.args.get('empresa_id', type=int)
        fecha_inicio = request.args.get('fecha_inicio')
        fecha_fin = request.args.get('fecha_fin')
        
        if not fecha_inicio or not fecha_fin:
            semana_offset = request.args.get('semana', type=int, default=0)
            hoy = datetime.now().date()
            inicio_semana = hoy - timedelta(days=hoy.weekday()) + timedelta(weeks=semana_offset)
            fin_semana = inicio_semana + timedelta(days=6)
            fecha_inicio = inicio_semana.strftime('%Y-%m-%d')
            fecha_fin = fin_semana.strftime('%Y-%m-%d')
        
        conn = get_connection()
        cursor = conn.cursor()
        
        empresa_filter = "AND e.empresa_id = %s" if empresa_id else ""
        params = [fecha_inicio, fecha_fin]
        if empresa_id:
            params.append(empresa_id)
        
        cursor.execute(f"""
            SELECT 
                e.nombre,
                (
                    COUNT(CASE 
                        WHEN a.entrada_manana_real IS NOT NULL 
                             AND TIME(a.entrada_manana_real) > '06:50:59'
                        THEN 1 
                    END) +
                    COUNT(CASE 
                        WHEN a.entrada_tarde_real IS NOT NULL 
                             AND TIME(a.entrada_tarde_real) > '14:50:59'
                        THEN 1 
                    END)
                ) as tardanzas,
                (
                    COUNT(CASE WHEN a.entrada_manana_real IS NOT NULL THEN 1 END) +
                    COUNT(CASE WHEN a.entrada_tarde_real IS NOT NULL THEN 1 END)
                ) as total_turnos
            FROM EMPLEADOS e
            INNER JOIN ASISTENCIA a ON e.id = a.empleado_id 
                AND a.fecha BETWEEN %s AND %s
            WHERE e.activo = TRUE {empresa_filter}
            GROUP BY e.id, e.nombre
            HAVING tardanzas > 0  -- 🎯 Al menos 1 tardanza
            ORDER BY tardanzas DESC, total_turnos DESC
            LIMIT 5
        """, params)
        
        top_tardes = []
        for row in cursor.fetchall():
            top_tardes.append({
                "nombre": row[0],
                "tardanzas": row[1],
                "total_turnos": row[2]
            })
        
        cursor.close()
        conn.close()
        
        return jsonify(top_tardes)
        
    except Exception as e:
        import traceback
        print(f"❌ Error en top late: {e}")
        print(traceback.format_exc())
        return jsonify({"error": str(e)}), 500
#Rutasa para ver empleados que no marcaron completo o se olvidaron
# Agregar estas rutas al final de app.py (antes del if __name__ == '__main__':)

@app.route('/admin/incomplete-markings')
def admin_incomplete_markings():
    """Página para gestionar marcaciones incompletas"""
    if not session.get('admin_logged_in'):
        flash('Debes iniciar sesión para acceder a esta sección', 'error')
        return redirect(url_for('admin_login'))
    
    empresas = list_companies_use_case.execute()
    return render_template('incomplete_markings.html', empresas=empresas)


@app.route('/api/incomplete-markings')
def api_incomplete_markings():
    """Obtiene registros de asistencia incompletos"""
    if not session.get('admin_logged_in'):
        return jsonify({"error": "No autorizado"}), 401
    
    try:
        from src.infrastructure.mysql_connection import get_connection
        
        empresa_id = request.args.get('empresa_id', type=int)
        mes = request.args.get('mes', type=int, default=datetime.now().month)
        anio = request.args.get('anio', type=int, default=datetime.now().year)
        
        conn = get_connection()
        cursor = conn.cursor()
        
        empresa_filter = "AND emp.id = %s" if empresa_id else ""
        params = [anio, mes]
        if empresa_id:
            params.append(empresa_id)
        
        # Buscar registros donde hay entrada pero no salida
        cursor.execute(f"""
            SELECT 
                a.id,
                a.empleado_id,
                e.nombre as empleado_nombre,
                emp.nombre as empresa_nombre,
                a.fecha,
                a.entrada_manana_real,
                a.salida_manana_real,
                a.entrada_tarde_real,
                a.salida_tarde_real,
                CASE 
                    WHEN a.entrada_manana_real IS NOT NULL AND a.salida_manana_real IS NULL THEN 'mañana'
                    WHEN a.entrada_tarde_real IS NOT NULL AND a.salida_tarde_real IS NULL THEN 'tarde'
                    WHEN a.entrada_manana_real IS NOT NULL AND a.salida_manana_real IS NULL 
                         AND a.entrada_tarde_real IS NOT NULL AND a.salida_tarde_real IS NULL THEN 'ambos'
                END as turno_incompleto
            FROM ASISTENCIA a
            JOIN EMPLEADOS e ON a.empleado_id = e.id
            JOIN EMPRESAS emp ON e.empresa_id = emp.id
            WHERE YEAR(a.fecha) = %s 
              AND MONTH(a.fecha) = %s
              AND e.activo = TRUE
              AND (
                  (a.entrada_manana_real IS NOT NULL AND a.salida_manana_real IS NULL) OR
                  (a.entrada_tarde_real IS NOT NULL AND a.salida_tarde_real IS NULL)
              )
              {empresa_filter}
            ORDER BY a.fecha DESC, e.nombre ASC
        """, params)
        
        registros = []
        for row in cursor.fetchall():
            turno_incompleto = row[9]
            turnos_faltantes = []
            
            if turno_incompleto in ['mañana', 'ambos']:
                turnos_faltantes.append({
                    'turno': 'mañana',
                    'entrada': str(row[5]) if row[5] else None,
                    'salida': None
                })
            
            if turno_incompleto in ['tarde', 'ambos']:
                turnos_faltantes.append({
                    'turno': 'tarde',
                    'entrada': str(row[7]) if row[7] else None,
                    'salida': None
                })
            
            registros.append({
                'asistencia_id': row[0],
                'empleado_id': row[1],
                'empleado_nombre': row[2],
                'empresa_nombre': row[3],
                'fecha': row[4].strftime('%Y-%m-%d'),
                'fecha_formato': row[4].strftime('%d/%m/%Y'),
                'dia_semana': row[4].strftime('%A'),
                'turnos_faltantes': turnos_faltantes
            })
        
        cursor.close()
        conn.close()
        
        # Traducir días de la semana
        dias_traduccion = {
            'Monday': 'Lunes', 'Tuesday': 'Martes', 'Wednesday': 'Miércoles',
            'Thursday': 'Jueves', 'Friday': 'Viernes', 'Saturday': 'Sábado', 'Sunday': 'Domingo'
        }
        
        for registro in registros:
            registro['dia_semana'] = dias_traduccion.get(registro['dia_semana'], registro['dia_semana'])
        
        return jsonify({
            'total': len(registros),
            'registros': registros
        })
        
    except Exception as e:
        import traceback
        print(f"❌ Error en incomplete-markings: {e}")
        print(traceback.format_exc())
        return jsonify({"error": str(e)}), 500


@app.route('/api/add-exit-time', methods=['POST'])
def api_add_exit_time():
    """Agregar hora de salida manualmente"""
    if not session.get('admin_logged_in'):
        return jsonify({"success": False, "message": "No autorizado"}), 401
    
    try:
        from src.infrastructure.mysql_connection import get_connection
        
        data = request.get_json()
        asistencia_id = data.get('asistencia_id')
        turno = data.get('turno')  # 'mañana' o 'tarde'
        hora_salida = data.get('hora_salida')  # formato "HH:MM"
        
        if not all([asistencia_id, turno, hora_salida]):
            return jsonify({"success": False, "message": "Datos incompletos"}), 400
        
        conn = get_connection()
        cursor = conn.cursor()
        
        # Obtener la fecha y hora de entrada del registro
        cursor.execute("""
            SELECT fecha, entrada_manana_real, entrada_tarde_real 
            FROM ASISTENCIA 
            WHERE id = %s
        """, (asistencia_id,))
        
        result = cursor.fetchone()
        if not result:
            cursor.close()
            conn.close()
            return jsonify({"success": False, "message": "Registro no encontrado"}), 404
        
        fecha = result[0]
        entrada_manana = result[1]
        entrada_tarde = result[2]
        
        # Validar que exista la entrada correspondiente
        if turno == 'mañana' and not entrada_manana:
            cursor.close()
            conn.close()
            return jsonify({"success": False, "message": "No existe entrada de mañana"}), 400
        
        if turno == 'tarde' and not entrada_tarde:
            cursor.close()
            conn.close()
            return jsonify({"success": False, "message": "No existe entrada de tarde"}), 400
        
        # Construir datetime completo
        fecha_str = fecha.strftime('%Y-%m-%d')
        salida_completa = f"{fecha_str} {hora_salida}:00"
        
        # Actualizar según el turno
        if turno == 'mañana':
            cursor.execute("""
                UPDATE ASISTENCIA 
                SET salida_manana_real = %s
                WHERE id = %s
            """, (salida_completa, asistencia_id))
        else:  # tarde
            cursor.execute("""
                UPDATE ASISTENCIA 
                SET salida_tarde_real = %s
                WHERE id = %s
            """, (salida_completa, asistencia_id))
        
        # Recalcular horas trabajadas y extras
        cursor.execute("""
            SELECT 
                entrada_manana_real, salida_manana_real,
                entrada_tarde_real, salida_tarde_real
            FROM ASISTENCIA WHERE id = %s
        """, (asistencia_id,))
        
        row = cursor.fetchone()
        entrada_m, salida_m, entrada_t, salida_t = row
        
        total_minutos = 0
        
        if entrada_m and salida_m:
            delta = salida_m - entrada_m
            total_minutos += int(delta.total_seconds() / 60)
        
        if entrada_t and salida_t:
            delta = salida_t - entrada_t
            total_minutos += int(delta.total_seconds() / 60)
        
        horas_trabajadas = total_minutos / 60.0
        horas_extras = max(0, horas_trabajadas - 8)
        
        cursor.execute("""
            UPDATE ASISTENCIA 
            SET horas_trabajadas = %s, horas_extras = %s
            WHERE id = %s
        """, (horas_trabajadas, horas_extras, asistencia_id))
        
        conn.commit()
        cursor.close()
        conn.close()
        
        return jsonify({
            "success": True, 
            "message": f"Salida de {turno} registrada correctamente",
            "horas_trabajadas": round(horas_trabajadas, 2),
            "horas_extras": round(horas_extras, 2)
        })
        
    except Exception as e:
        import traceback
        print(f"❌ Error agregando salida: {e}")
        print(traceback.format_exc())
        return jsonify({"success": False, "message": str(e)}), 500


@app.route('/api/delete-incomplete-marking', methods=['POST'])
def api_delete_incomplete_marking():
    """Eliminar registro incompleto"""
    if not session.get('admin_logged_in'):
        return jsonify({"success": False, "message": "No autorizado"}), 401
    
    try:
        from src.infrastructure.mysql_connection import get_connection
        
        data = request.get_json()
        asistencia_id = data.get('asistencia_id')
        
        if not asistencia_id:
            return jsonify({"success": False, "message": "ID requerido"}), 400
        
        conn = get_connection()
        cursor = conn.cursor()
        
        cursor.execute("DELETE FROM ASISTENCIA WHERE id = %s", (asistencia_id,))
        conn.commit()
        
        cursor.close()
        conn.close()
        
        return jsonify({
            "success": True, 
            "message": "Registro eliminado correctamente"
        })
        
    except Exception as e:
        import traceback
        print(f"❌ Error eliminando registro: {e}")
        print(traceback.format_exc())
        return jsonify({"success": False, "message": str(e)}), 500

# GESTIÓN DE REGISTROS DE ASISTENCIA PRO..

@app.route('/admin/attendance-records')
def admin_attendance_records():
    """Página principal de gestión de registros de asistencia"""
    if not session.get('admin_logged_in'):
        flash('Debes iniciar sesión para acceder a esta sección', 'error')
        return redirect(url_for('admin_login'))
    
    empresas = list_companies_use_case.execute()
    return render_template('attendance_records.html', empresas=empresas)


@app.route('/api/attendance-records')
def api_attendance_records():
    """Obtiene registros de asistencia con filtros"""
    if not session.get('admin_logged_in'):
        return jsonify({"error": "No autorizado"}), 401
    
    try:
        from src.infrastructure.mysql_connection import get_connection
        
        empresa_id = request.args.get('empresa_id', type=int)
        empleado_id = request.args.get('empleado_id', type=int)
        mes = request.args.get('mes', type=int, default=datetime.now().month)
        anio = request.args.get('anio', type=int, default=datetime.now().year)
        
        if not empresa_id:
            return jsonify({"error": "Empresa ID requerido"}), 400
        
        conn = get_connection()
        cursor = conn.cursor()
        
        # Query simplificado - Solo las horas de entrada y salida
        query = """
            SELECT 
                a.id as asistencia_id,
                a.empleado_id,
                e.nombre as empleado_nombre,
                emp.nombre as empresa_nombre,
                a.fecha,
                a.entrada_manana_real,
                a.salida_manana_real,
                a.entrada_tarde_real,
                a.salida_tarde_real
            FROM ASISTENCIA a
            JOIN EMPLEADOS e ON a.empleado_id = e.id
            JOIN EMPRESAS emp ON e.empresa_id = emp.id
            WHERE YEAR(a.fecha) = %s 
              AND MONTH(a.fecha) = %s
              AND emp.id = %s
        """
        
        params = [anio, mes, empresa_id]
        
        if empleado_id:
            query += " AND e.id = %s"
            params.append(empleado_id)
        
        query += " ORDER BY a.fecha DESC, e.nombre ASC"
        
        cursor.execute(query, params)
        
        registros = []
        dias_traduccion = {
            'Monday': 'Lunes', 'Tuesday': 'Martes', 'Wednesday': 'Miércoles',
            'Thursday': 'Jueves', 'Friday': 'Viernes', 'Saturday': 'Sábado', 'Sunday': 'Domingo'
        }
        
        for row in cursor.fetchall():
            fecha_obj = row[4]
            dia_semana_en = fecha_obj.strftime('%A')
            dia_semana_es = dias_traduccion.get(dia_semana_en, dia_semana_en)
            
            # Función auxiliar para formatear datetime a HH:MM
            def formatear_hora_bd(valor):
                if not valor:
                    return None
                if isinstance(valor, str):
                    # Si ya es string, extraer HH:MM
                    partes = valor.split(' ')
                    if len(partes) > 1:
                        return partes[1][:5]  # HH:MM
                    return valor[:5]
                elif hasattr(valor, 'strftime'):
                    # Si es datetime, formatear
                    return valor.strftime('%H:%M')
                elif hasattr(valor, 'total_seconds'):
                    # Si es timedelta
                    total_seconds = int(valor.total_seconds())
                    horas = total_seconds // 3600
                    minutos = (total_seconds % 3600) // 60
                    return f"{horas:02d}:{minutos:02d}"
                return str(valor)[:5] if valor else None
            
            registros.append({
                'asistencia_id': row[0],
                'empleado_id': row[1],
                'empleado_nombre': row[2],
                'empresa_nombre': row[3],
                'fecha': fecha_obj.strftime('%Y-%m-%d'),
                'fecha_formato': fecha_obj.strftime('%d/%m/%Y'),
                'dia_semana': dia_semana_es,
                'entrada_manana_real': formatear_hora_bd(row[5]),
                'salida_manana_real': formatear_hora_bd(row[6]),
                'entrada_tarde_real': formatear_hora_bd(row[7]),
                'salida_tarde_real': formatear_hora_bd(row[8])
            })
        
        cursor.close()
        conn.close()
        
        return jsonify({
            'total': len(registros),
            'registros': registros
        })
        
    except Exception as e:
        import traceback
        print(f"❌ Error en attendance-records: {e}")
        print(traceback.format_exc())
        return jsonify({"error": str(e)}), 500


@app.route('/api/update-attendance-record', methods=['POST'])
def api_update_attendance_record():
    """Actualizar registro de asistencia - Solo horas de entrada/salida"""
    if not session.get('admin_logged_in'):
        return jsonify({"success": False, "message": "No autorizado"}), 401
    
    try:
        from src.infrastructure.mysql_connection import get_connection
        
        data = request.get_json()
        asistencia_id = data.get('asistencia_id')
        entrada_manana = data.get('entrada_manana')
        salida_manana = data.get('salida_manana')
        entrada_tarde = data.get('entrada_tarde')
        salida_tarde = data.get('salida_tarde')
        
        if not asistencia_id:
            return jsonify({"success": False, "message": "ID de asistencia requerido"}), 400
        
        conn = get_connection()
        cursor = conn.cursor()
        
        # Obtener la fecha del registro
        cursor.execute("SELECT fecha FROM ASISTENCIA WHERE id = %s", (asistencia_id,))
        result = cursor.fetchone()
        
        if not result:
            cursor.close()
            conn.close()
            return jsonify({"success": False, "message": "Registro no encontrado"}), 404
        
        fecha = result[0]
        fecha_str = fecha.strftime('%Y-%m-%d')
        
        # Construir valores datetime completos (o NULL si está vacío)
        entrada_manana_dt = f"{fecha_str} {entrada_manana}:00" if entrada_manana else None
        salida_manana_dt = f"{fecha_str} {salida_manana}:00" if salida_manana else None
        entrada_tarde_dt = f"{fecha_str} {entrada_tarde}:00" if entrada_tarde else None
        salida_tarde_dt = f"{fecha_str} {salida_tarde}:00" if salida_tarde else None
        
        # Actualizar SOLO las horas de entrada/salida
        cursor.execute("""
            UPDATE ASISTENCIA 
            SET entrada_manana_real = %s,
                salida_manana_real = %s,
                entrada_tarde_real = %s,
                salida_tarde_real = %s
            WHERE id = %s
        """, (entrada_manana_dt, salida_manana_dt, entrada_tarde_dt, salida_tarde_dt, asistencia_id))
        
        conn.commit()
        cursor.close()
        conn.close()
        
        return jsonify({
            "success": True, 
            "message": "Horarios actualizados correctamente"
        })
        
    except Exception as e:
        import traceback
        print(f"❌ Error actualizando registro: {e}")
        print(traceback.format_exc())
        return jsonify({"success": False, "message": str(e)}), 500


@app.route('/api/delete-attendance-record', methods=['POST'])
def api_delete_attendance_record():
    """Eliminar registro de asistencia completo"""
    if not session.get('admin_logged_in'):
        return jsonify({"success": False, "message": "No autorizado"}), 401
    
    try:
        from src.infrastructure.mysql_connection import get_connection
        
        data = request.get_json()
        asistencia_id = data.get('asistencia_id')
        
        if not asistencia_id:
            return jsonify({"success": False, "message": "ID requerido"}), 400
        
        conn = get_connection()
        cursor = conn.cursor()
        
        # Verificar que existe
        cursor.execute("SELECT id FROM ASISTENCIA WHERE id = %s", (asistencia_id,))
        if not cursor.fetchone():
            cursor.close()
            conn.close()
            return jsonify({"success": False, "message": "Registro no encontrado"}), 404
        
        # Eliminar
        cursor.execute("DELETE FROM ASISTENCIA WHERE id = %s", (asistencia_id,))
        conn.commit()
        
        cursor.close()
        conn.close()
        
        return jsonify({
            "success": True, 
            "message": "Registro eliminado correctamente"
        })
        
    except Exception as e:
        import traceback
        print(f"❌ Error eliminando registro: {e}")
        print(traceback.format_exc())
        return jsonify({"success": False, "message": str(e)}), 500
if __name__ == '__main__':
    app.run(debug=False, host='0.0.0.0', port=8080)